#start to build a working file

from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Fill
wb = load_workbook(filename='chart.xlsx')  # file to open must be in the same dirctory or must switch to get the file
ws1 = wb['Defects'] #selects the tab
ws1.sheet_properties.tabColor = "1072BA" #Colors must be aRGB hex values
a1 = ws1['A1']
a2 = ws1['A2']
a3 = ws1['A3']
a4 = ws1['A4']
ft1 = Font(name='Arial', size=22, color=colors.RED, bold=True)#If you want to apply styles to entire rows and columns then you must apply the style to each cell yourself. 
a1.font = ft1
a2.font = ft1
a3.font = ft1
a4.font = ft1#chnages cell A1 to red and 22 font

#ws1.conditional_formatting('A1:A10', FormulaRule(formula= ['A2'>= 99], fill=redFill) 
#col = ws1.column_dimensions['A']
#col.font = Font(bold=True)
#row = ws1.row_dimensions[10]
#row.font = Font(underline="single")
# openpyxl.worksheet.worksheet.Worksheet.insert_rows()
# openpyxl.worksheet.worksheet.Worksheet.insert_cols()
# openpyxl.worksheet.worksheet.Worksheet.delete_rows()
# openpyxl.worksheet.worksheet.Worksheet.delete_cols()

ws1.insert_cols(5, 9)
ws1.delete_cols(6, 3)#deletes columns


print(ws1)

wb.save('New_Chart2.xlsx')
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended - adds the values
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

