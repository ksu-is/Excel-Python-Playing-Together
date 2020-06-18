

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

a1 = ws['A1']
d4 = ws['D4']
#ft = Font(color=colors.RED)
ft1 = Font(name='Arial', size=14, color=colors.RED)
#a1.font = ft
d4.font = ft1
a1.font = ft1

#ft2 = copy(ft1)
#ft2.name = "Tahoma"
#ft1.name

#ft2.name

#ft2.size





wb.save('testing.xlsx')

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended - adds the values
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")