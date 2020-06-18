#Python code to Copy data from one excel file to another

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.colors import YELLOW, BLUE, GREEN, RED
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles.alignment import Alignment
import datetime
from openpyxl.utils import FORMULAE


src_wb = load_workbook('AMMM_OTD_30_60_90.xlsx') #opens data source file
src_sheet = src_wb['Sheet1'] #selects the active sheet specified by name

wb = Workbook() #starts creation of ne workbook
ws = wb.active
ws.title = "AMMM OTD 306090"

for i in range(1, src_sheet.max_row+1): #captures data from source file and pastes to new worksheet
    for j in range(1, src_sheet.max_column+1): #this iterates through all the data to the end of each set, not the entires sheet
        ws.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value
        ws.cell(row=i, column=j).font = Font(name='Arial', size=9, bold=True)#basic format for data pasted

ws.delete_cols(1) #deletes blank rows/columns from source
ws.delete_rows(1)
#formattiong the header row coumn headings
for col_range in range(1, 6):
    cell_title = ws.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="1B289F", end_color="1B289F", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='FAFBFD')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(6, 12):
    cell_title = ws.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="0EF0F7", end_color="0EF0F7", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(12, 17):
    cell_title = ws.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(17, 25):
    cell_title = ws.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="1B289F", end_color="1B289F", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='FAFBFD')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(25, 28):
    cell_title = ws.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)

#formats particular column data with either number or date 
for col in ws.iter_cols(min_row=2, min_col= 11, max_col=12, max_row=1001):#start with Row 2 as 1 is the header row
     for cell in col:
        cell.number_format = 'mm/dd/yyyy'
for col in ws.iter_cols(min_row=2, min_col= 25, max_col=26, max_row=1001):
     for cell in col:
        cell.number_format = 'mm/dd/yyyy'
for col in ws.iter_cols(min_row=2, min_col= 17, max_col=17, max_row=1001):
     for cell in col:
        cell.number_format = '#####0'        

# captures the data converts it to a string then concatentes the string into the column rows 
for item in range(2, ws.max_row+1):
    ws.cell(row=item, column=27).value = str(ws.cell(row=item, column=7).value) + str(ws.cell(row=item, column=8).value) + str(ws.cell(row=item, column=9).value) + str(ws.cell(row=item, column=11).value)

# Conditional Formatting: Highlights cells that contain particular text by using a special formula
#Red text for 'LATE'
red_text = Font(color="000000", bold=True)
red_fill = PatternFill(bgColor="FF0000")
dxf1 = DifferentialStyle(font=red_text, fill=red_fill)
rule1 = Rule(type="containsText", operator="containsText", text="LATE", dxf=dxf1)
rule1.formula = ['NOT(ISERROR(SEARCH("LATE",N1)))']
ws.conditional_formatting.add('I1:I1001', rule1)
#Yellow Text for 'ON TIME'
y_text = Font(color="000000", bold=True)
y_fill = PatternFill(bgColor="FFFF00")
dxf2 = DifferentialStyle(font=y_text, fill=y_fill)
rule2 = Rule(type="containsText", operator="containsText", text="ON TIME", dxf=dxf2)
rule2.formula = ['NOT(ISERROR(SEARCH("ON TIME",N1)))']
ws.conditional_formatting.add('I1:I1001', rule2)
#Orange fo 'Contract Change'
o_text = Font(color="000000", bold=True)
o_fill = PatternFill(bgColor="FF8000")
dxf3 = DifferentialStyle(font=o_text, fill=o_fill)
rule3 = Rule(type="containsText", operator="containsText", text="Contract", dxf=dxf3)
rule3.formula = ['NOT(ISERROR(SEARCH("Contract",O1)))']
ws.conditional_formatting.add('I1:I1001', rule3)
#Green for 'DD250'
g_text = Font(color="000000", bold=True)
g_fill = PatternFill(bgColor="00FF00")
dxf4 = DifferentialStyle(font=g_text, fill=g_fill)
rule4 = Rule(type="containsText", operator="containsText", text="DD250", dxf=dxf4)
rule4.formula = ['NOT(ISERROR(SEARCH("DD250",O1)))']
ws.conditional_formatting.add('I1:I1001', rule4)


ws['AA1'] = ('concat')#adds text to a particular cell
# set the width of the column 
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 15
ws.column_dimensions['O'].width = 23
ws.column_dimensions['P'].width = 62
ws.column_dimensions['Y'].width = 12
ws.column_dimensions['Z'].width = 12
ws.auto_filter.ref = "A1:AA1" #adds filters


src_wb.save('AMMM_OTD_30_60_90.xlsx')

wb.save('new_excelfile4.xlsx') # saves as new file