#Python code to Copy data from one excel file to another

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.colors import YELLOW, BLUE, GREEN, RED
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.alignment import Alignment

src_wb = load_workbook('AMMM_OTD_30_60_90.xlsx')
dest_wb = load_workbook('new_excelfile.xlsx')

src_sheet = src_wb['Sheet1']
dest_sheet = dest_wb['range names']
ws = dest_wb.active

#testing for new workbook
wb = Workbook()

ws = wb.active
ws.title = "AMMM OTD 306090"
for i in range(1, src_sheet.max_row+1): 
    for j in range(1, src_sheet.max_column+1):
        ws.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value
        ws.cell(row=i, column=j).font = Font(name='Arial', size=9, bold=True)
#this worked to put data into a new workbook


for i in range(1, src_sheet.max_row+1): 
    for j in range(1, src_sheet.max_column+1):
        dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value
        dest_sheet.cell(row=i, column=j).font = Font(name='Arial', size=9, bold=True)
ws.delete_cols(1)
ws.delete_rows(1)
for col_range in range(1, 6):
    cell_title = dest_sheet.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="1B289F", end_color="1B289F", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='FAFBFD')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(6, 12):
    cell_title = dest_sheet.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="0EF0F7", end_color="0EF0F7", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(12, 17):
    cell_title = dest_sheet.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(17, 25):
    cell_title = dest_sheet.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="1B289F", end_color="1B289F", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='FAFBFD')
    cell_title.alignment = Alignment(wrap_text=True)
for col_range in range(25, 28):
    cell_title = dest_sheet.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell_title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    cell_title.alignment = Alignment(wrap_text=True)

ws['A1'] = ('MCS ACO Code')
src_sheet['A27'] = ('concat')

src_wb.save('AMMM_OTD_30_60_90.xlsx')
dest_wb.save('new_excelfile.xlsx')
wb.save('new_excelfile4.xlsx')