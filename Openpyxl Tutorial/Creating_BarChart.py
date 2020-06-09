#a bar chart

#from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='chart.xlsx')
#wb = Workbook()
ws = wb.active
for i in range(10):
    ws.append([i])

from openpyxl.chart import BarChart, Reference, Series
values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")
wb.save("SampleBarChart.xlsx")