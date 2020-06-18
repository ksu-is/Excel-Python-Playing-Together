
# write a workbook from scratch

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))#filled spreadsheet rows with numbers 1-599

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws2['F6'] = 3.14
ws2['F7'] = 3.14
ws2['F8'] = 3.14
ws2['F9'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):#puts data startinf at row 10 column AA
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save('new_excelfile.xlsx')

