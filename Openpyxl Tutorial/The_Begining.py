# the beginning

from openpyxl import Workbook

wb = Workbook()  # There is no need to create a file on the filesystem to get started with openpyxl.
# Just import the Workbook class and start work:

ws = wb.active  # A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:
ws1 = wb.create_sheet("Mysheet")  # insert at the end (default)
# or
ws2 = wb.create_sheet("Yoursheet", 0)  # insert at first position
# or
ws3 = wb.create_sheet("Thesheet", -1)  # insert at the penultimate position

ws.title = "New Title"  # gives a new title
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]  # you can get it as a key of the workbook:
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)  # create copies of worksheets within a single workbook:

# accessing cells and values
c = ws['A4']  # creates or adds a cell
c = ws['A4'] = 4
# there is a method
d = ws.cell(row=4, column=2, value=10)
# cells can be accesses by slicing
cell_range = ws['A1':'C2']

# ranges of rows or columns can be obtained
colC = ws['C']
col_range = ws["C:D"]
row10 = ws[10]
row_range = ws[5:10]

# use the Worksheet.iter_rows() method: for rows
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# use the Worksheet.iter_cols() method: for columns
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
# For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.

ws = wb.active
ws['C9'] = 'hello world'
mo = tuple(ws.rows)
print()
print(mo)
# or columns
mo = tuple(ws.columns)
print()
print(mo)
print()

# Values - if just the values use the Worksheet.values property.
for row in ws.values:
    for value in row:
        print(value)
print("break")

# Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# we can assign values to the cells
c = ws['B2']
c.value = 'hello, world'
print(c.value)

d = ws['A2']
d.value = 3.14
print(d.value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

#wb = Workbook()
wb.save('balances.xlsx')

#from openpyxl import load_workbook

#wb2 = load_workbook('chart.xlsx')  # file to open must be in the same dirctory or must switch to get the file
#print(wb2.sheetnames)
