# Using number formats
import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
# set date using a Python datetime
ws['A1'] = datetime.datetime(2010, 7, 21)

print(ws['A1'].number_format)

# Using formulae

wb = Workbook()
ws = wb.active
# add a simple formula
ws["A1"] = "=SUM(1, 1)"
wb.save("formula.xlsx")

# merging cells
# from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

ws.merge_cells('A2:D2')
# ws.unmerge_cells('A2:D2')

# or equivalently
ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=4)
# ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
wb.save("merged.xlsx")
