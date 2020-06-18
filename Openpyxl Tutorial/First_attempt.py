# first attempt to create a file from data


from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

wb = Workbook()
# grab the active worksheet
ws = wb.active
ws.title = "AMMM OTD 306090"
for x in range(1,1001):
       for y in range(1,1001):
           ws.cell(row=x, column=y)
cell_range = ws['A1':'A40']
#for row in range(1, 40):
    #ws.append(['MCS ACO Code', 'ECS LOB', 'Contract Number', 'Delivery Order', 'GK Platform', 'MCS ISP Code', 'MCS PIIN', 'MCS SPIIN', 'MCS Line Item ID', 'MCS Noun', 'MCS Delivery Schedule Date', 'WB ECD', 
   # 'GK CLIN Program Manager', 'High-Level Delivery Status', 'WB Status', 'WB Comments', 'MCS Bal Qty Shed', 'WB CLIN Status POC', 'GK LM Contracts POC', 'ECS CLIN Description', 'MCS Part Number', 'MCS Current Status',
   # 'MCS Contract Purpose', 'MCS Contract Type', 'MOCAS Delivery Due Date', 'MCS Service Completion Date', 'concat'])



# Data can be assigned directly to cells
ws['A1'] = ('MCS ACO Code')

# Rows can also be appended - adds the values
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()





wb.save('new_excelfile2.xlsx')
