from openpyxl import load_workbook
workbook = load_workbook(filename="My_Test_Data.xlsx")

print("Names are ",workbook.sheetnames)
sheet = workbook.active
print("Sheet type is: ",type(sheet))


sheet.title
print(sheet.title)