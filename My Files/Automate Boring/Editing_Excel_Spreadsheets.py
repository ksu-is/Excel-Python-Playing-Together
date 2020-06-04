#Editing Excel Spreadsheets

import openpyxl

wb= openpyxl.load_workbook('chart.xlsx')#opens workbook
sheet = wb.get_sheet_by_name('Sheet1')#goes to specific sheet
print(sheet['B1'].value)# get data from specific cell as a value

#to create a new workbook
wb=openpyxl.Workbook()
sheet=wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value)

sheet['A1'] = 42
sheet['A2'] = "hello World!"

wb.save('Editing Example2.xlsx')#created a new spreadsheet file

wb= openpyxl.load_workbook('chart.xlsx')#opens workbook
wb.save('chart2.xlsx')#created a new spreadsheet file and retain original file
sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
sheet2.title=('My New Sheet Name!')#change name of a sheet
print(wb.get_sheet_names())
wb.save('chart3.xlsx')
wb.create_sheet(index=1, title='My Other New Spreadsheet')
wb.save('chart3.xlsx')