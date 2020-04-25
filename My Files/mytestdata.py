from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.axis import DateAxis
    #Line 7  how to create a new empty workbook.
    #Lines 13 & 14  how to add data to specific cells.
    #Line 17 how to save the spreadsheet 

filename = "My_Test_Data.xlsx"

#workbook = Workbook()

#sheet = workbook.active

workbook = load_workbook(filename="My_Test_Data.xlsx")
sheet = workbook.active

#this is a bar chart
# chart = BarChart()

#data = Reference(worksheet=sheet,
                # min_row=1,
                 #max_row=8,
                 #min_col=2,
                 #max_col=3)

#chart.add_data(data, titles_from_data=True)

#sheet.add_chart(chart, "E2")
chart = LineChart()
data = Reference(worksheet=sheet,
                 min_row=2,
                 max_row=23,
                 min_col=1,
                 max_col=70)
chart.add_data(data, from_rows=True, titles_from_data=True)

sheet.add_chart(chart, "C6")
chart.x_axis.title = "Weeks"
chart.y_axis.title = "Count"
chart.title= " Count per Aircraft "
chart.y_axis.title = "Count"
#legend.position = "l"
fill = PatternFill(fill_type=None,
                start_color='FFFFFFFF',
                end_color='FF000000')
workbook.save("line_chart.xlsx")

workbook.save("barchart.xlsx")