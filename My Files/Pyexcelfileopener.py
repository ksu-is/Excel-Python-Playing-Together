from openpyxl import load_workbook
workbook = load_workbook(filename="reviews-sample1.xlsx")

print("Names are ",workbook.sheetnames)
#print("workbook type is", type(workbook))
sheet = workbook.active
#print("Sheet type is: ",type(sheet))


sheet.title
print(sheet.title)
sheet["A1"].value

#print(sheet["A1"].value)
#print(sheet["F10"].value)
#print(sheet.cell(row=10, column=6).value)
#print(sheet["A1:C2"])
#for row in sheet.iter_rows(min_row=1,
                           #max_row=2,
                           #min_col=1,
                           #max_col=3,
                           #values_only=True):
    #print(row)
#testing for different values to see different results from sample data
#for value in sheet.iter_rows(min_row=2,
                             #min_col=4,
                             #max_col=7,
                             #values_only=True):
    #print(value)

#for value in sheet.iter_rows(min_row=2,
                            # min_col=4,
                             #max_col=8,
                             #values_only=True):
    #print(value)

# shows the names of the column headings
#for value in sheet.iter_rows(min_row=1,
                             #max_row=1,
                             #values_only=True):
    #print(value)
 #shows values of selected columns 
#for value in sheet.iter_rows(min_row=2,
                             # min_col=4,
                              #max_col=7,
                              #values_only=True):
    #print(value)
import json
from openpyxl import load_workbook

workbook = load_workbook(filename="reviews-sample1.xlsx")
sheet = workbook.active

products = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(products))

import datetime
from dataclasses import dataclass

@dataclass
class Product:
    id: str
    parent: str
    title: str
    category: str

@dataclass
class Review:
    id: str
    customer_id: str
    stars: int
    headline: str
    body: str
    date: datetime.datetime

for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)


# Or an alternative
for cell in sheet[1]:
    print(cell.value)
