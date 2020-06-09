#Automate_Excel

import openpyxl

workbook = openpyxl.load_workbook('Automate_Excel_Example.xlsx')
type(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')
type(sheet)

mo = workbook.get_sheet_names()
print(mo)

sheet['A1']
cell = sheet['A1']
print(cell.value)

str(cell.value)

str(sheet['A1'])


print(sheet['B1'].value)
print(sheet['C1'].value)

mo = sheet.cell(row=1, column=2).value #returns the value of the cell
print(mo)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)# prints all data in column B
