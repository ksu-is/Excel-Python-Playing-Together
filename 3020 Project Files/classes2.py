import datetime
from dataclasses import dataclass
import json
from openpyxl import load_workbook
workbook = load_workbook(filename="reviews-sample1.xlsx")
sheet = workbook.active

@dataclass
class Product:
    id: str
    parent: str
    title: str
    category: str

@dataclass
class Review:
    id: str
    customer_id: str
    stars: int
    headline: str
    body: str
    date: datetime.datetime

for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)


# Or an alternative
for cell in sheet[1]:
    print(cell.value)

